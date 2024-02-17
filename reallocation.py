
from collections import Counter, OrderedDict

import openpyxl

from common import STORE_MAP, PASS, SUCCESS, JINNY_THINK, JOY_CONFIRM, NO_6M_SALES
from common import get_current_time, get_column_number, get_last_data_idx, convert_to_zero


SOURCE_DIR = "C:/Users/jinny.hur/Desktop/업무파일/01. 리얼로케이션" # 마스터 파일 저장 경로
SOURCE_FILE_NAME = "Retail BTQ reallocation (0207_rank)" # 마스터 파일명(확장자 제외)

TARGET_ROTATION = 1.5


def get_retail_info_template(stock_info_list, store_info_list):
    store_keys_list = list(STORE_MAP.keys())

    stock_info = {
        "total_available": stock_info_list[0], 
        "tobe_moved": 0,
        "warehouse": {
            "KRD4/1001": stock_info_list[4],
            "from_KRD4/1001": 0
        },
        "shipping": {
            f"shipment{i}": convert_to_zero(stock_info_list[(i-1) * 2 + 8]) for i in range(1, 4)
        }
    }
    stock_info["shipping"].update({
        f"from_shipment{i}": convert_to_zero(stock_info_list[(i-1) * 2 + 9]) for i in range(1, 4)
    })

    store_info = {
        store_keys_list[i]: {
            "L12M_sales": convert_to_zero(store_info_list[i * 10]),
            "L6M_sales": convert_to_zero(store_info_list[i * 10 + 1]),
            "L3M_sales": convert_to_zero(store_info_list[i * 10 + 2]),
            "MS": convert_to_zero(store_info_list[i * 10 + 3]),
            "stock": convert_to_zero(store_info_list[i * 10 + 4]),
            "in_transit": convert_to_zero(store_info_list[i * 10 + 5]),
            "wish_list": convert_to_zero(store_info_list[i * 10 + 6]),
            "reallocation": convert_to_zero(store_info_list[i * 10 + 7]),
            "coverage": convert_to_zero(store_info_list[i * 10 + 8]),
            "rotation": store_info_list[i * 10 + 9], # rotation의 경우 수식값으로 None과 0.0 이 명확하게 구분되는 바 filter 적용하지 않음
        } for i in range(len(store_keys_list))
    }

    return {"result": "", "stock_info": stock_info, "store_info": store_info}


class Main:
    def __init__(self):
        pass
    
    def run(self):
        """
            1. 데이터 조회
                : 아래 data-set 양식으로 마스터 파일 변환
            2. reallocation
                a. store_info('24.2.11)
                    1. total_available 값이 없으면 pass
                    2. rotation 값 기준으로 지점 별 내림차순하여, 최저값 지점에 1개씩 제공
                    3. 각 재고 배부 이후 다시 전체 rotation 조회하며, 2번 반복
                    4. 단, 3번 과정 중, 현재 재고로 최저 rotation 지점 배분 불가 시(최저 rotation 지점 개수 > 현재 재고) JINNY_THINK 분류

                b. stock_info('24.2.11)
                    1. TTL Available 값에 대하여 KRD4/1001, Shipment 1, Shipment 2, Shipment 3 순으로 배분
                    2. 단, success로 판정되지 않았으면 stock_info를 조정하지 않음 (결국 store_info부터 stock_info까지 JINNY_THINK 필요)
                    3. 또한, KRD4_1001 재고 사용 필요 시, JOY_CONFIRM 로 분류

            3. 엑셀 반환

            # data_set
            {
                "CRN7413800": {
                    "result": PASS / SUCCESS / JINNY_THINK / JOY_CONFIRM
                    "stock_info": {
                        "total_available": a
                        "tobe_moved": b,
                        "warehouse": {
                            "KRD4/1001": c
                        },
                        "shipping": {
                            "shipment1": d,
                            "shipment2": e,
                            "shipment3": f
                        }
                    },
                    "store_info": {
                        "6400: {
                            "L12M_sales": a,
                            "L6M_sales": a,
                            "L3M sales": b,
                            "MS": c,
                            "stock": d,
                            "in_transit": e,
                            "wish_list": f,
                            "reallocation": g,
                            "coverage": h,
                            "rotation": i,
                        },
                        ...
                    }
                },
                ...
            }
        """

        # 마스터 파일의 "Allocation" 시트를 참조하여, 위 형태의 data-set을 구성
        file_name = f'{SOURCE_DIR}/{SOURCE_FILE_NAME}.xlsx'
        print(f"Reading {file_name} ...")
        wb = openpyxl.load_workbook(filename=file_name, data_only=True)
        ws = wb["Allocation"]

        retail_info = dict()
        for row in ws.iter_rows(min_row=8, max_row=get_last_data_idx(ws), values_only=True):
            temp = get_retail_info_template(list(row)[get_column_number("N"):get_column_number("AA") + 1] # stock_info
                                          , list(row)[get_column_number("AK"):get_column_number("EP") + 1]) # store_info
            retail_info[row[1]] = temp

        # reallocation for store_info
        print("----- STORE_INFO REALLOCATION -----\n")
        for ref_no, retail in retail_info.items():
            # for debugging
            # if ref_no not in ['CRB4086444', 'CRB4086445']:
            #     continue

            total_available = retail.get("stock_info").get("total_available")
            current_available = total_available # reallocation 위해 재고값 복사

            print(f'\n# START {ref_no} with {total_available} stock')

            if total_available == 0.0:
                print(f'{ref_no} pass since no TTL Available')
                retail_info[ref_no]["result"] = PASS
                continue

            while True:
                if current_available == 0:
                    print(f'### store reallocation finished\n')
                    break

                # 유효한 모든 지점의 rotation 값이 TARGET_ROTATION 초과 시, SUCCESS
                if all(item[1].get("rotation") > TARGET_ROTATION for item in retail["store_info"].items() if item[1].get("rotation") is not None):
                    print(f'### {SUCCESS}\n- all store rotation over {TARGET_ROTATION} or all rotation None\n')
                    retail_info[ref_no]["result"] = SUCCESS
                    break

                # L6M_sales 값이 존재하는 store 정보만 가져온다. (엑셀에서도 없을 시 제외하고 있음)
                store_info_items = [item for item in retail["store_info"].items() if item[1].get("L6M_sales") > 0.0]
                if len(store_info_items) == 0: # 전 지점에서 6개월 판매값이 없을 시
                    print(f'### {NO_6M_SALES}\n- no L6M_sales in all stores\n')
                    retail_info[ref_no]["result"] = NO_6M_SALES
                    break

                # print(store_info_items)

                sorted_dict = OrderedDict(sorted(store_info_items, key=lambda x: x[1].get("rotation", 0) if isinstance(x[1], dict) else 0))
                rotation_counts = Counter(value.get("rotation", 0) if isinstance(value, dict) and isinstance(value.get("rotation"), (int, float)) else 0 for key, value in sorted_dict.items())
                rotation_counts = OrderedDict(sorted(rotation_counts.items()))

                lowest_rotation_count = rotation_counts[list(rotation_counts.keys())[0]]
                if lowest_rotation_count > current_available: # 현 재고에 대하여 더 이상 지점별로 동등하게 배분 못할 시
                    print(f'### {JINNY_THINK}\n- {rotation_counts} > {current_available}\n')
                    retail_info[ref_no]["result"] = JINNY_THINK
                    break

                sorted_list = sorted(store_info_items, key=lambda x: x[1].get("rotation", 0))
                store_id, store_info = sorted_list[0] # 가장 rotation이 낮은 store을 대상으로 함

                current_reallocation = store_info["reallocation"]
                current_rotation = store_info["rotation"]

                retail_info[ref_no]["store_info"][store_id]["reallocation"] += 1
                retail_info[ref_no]["store_info"][store_id]["rotation"] = (store_info["stock"] + store_info["in_transit"] + (current_reallocation + 1)) / (store_info["L6M_sales"] / 6)
                retail_info[ref_no]["result"] = SUCCESS
                retail_info[ref_no]["stock_info"]["tobe_moved"] += 1 # 실제 reallocation 대상 ++

                current_available -= 1 # reallocation 가능 재고 차감

                msg = f'## {STORE_MAP.get(store_id)}({store_id})\n'
                msg += f'- reallocation: {current_reallocation} → {retail_info[ref_no]["store_info"][store_id]["reallocation"]}\n'
                msg += f'- rotation: {current_rotation} → {retail_info[ref_no]["store_info"][store_id]["rotation"]}'
                
                print(msg)
                print("")
                # time.sleep(3)

        # statistics & reallocation for stock_info
        print("----- STOCK_INFO REALLOCATION -----\n")
        for ref_no, retail in retail_info.items():
            if retail.get("result") not in [SUCCESS]: # SUCCESS 아니면, 재고정보 변경하지 않는다 
                continue

            total_available = retail.get("stock_info").get("total_available")
            tobe_moved = retail.get("stock_info").get("tobe_moved")
            print(f'\n# START {ref_no} with {tobe_moved} target')

            while tobe_moved > 0:
                KRD4_1001, from_KRD4_1001 = convert_to_zero(retail["stock_info"]["warehouse"]["KRD4/1001"]), convert_to_zero(retail["stock_info"]["warehouse"]["from_KRD4/1001"])
                if KRD4_1001 > from_KRD4_1001:
                    print(f'## stock added on warehouse-from_KRD4/1001')
                    retail["stock_info"]["warehouse"]["from_KRD4/1001"] = from_KRD4_1001 + 1
                    tobe_moved -= 1
                    continue

                shipment1, from_shipment1 = convert_to_zero(retail["stock_info"]["shipping"]["shipment1"]), convert_to_zero(retail["stock_info"]["shipping"]["from_shipment1"])
                if shipment1 > from_shipment1:
                    print(f'## stock added on shipping-from_shipment1')
                    retail["stock_info"]["shipping"]["from_shipment1"] = from_shipment1 + 1
                    tobe_moved -= 1
                    continue
                
                shipment2, from_shipment2 = convert_to_zero(retail["stock_info"]["shipping"]["shipment2"]), convert_to_zero(retail["stock_info"]["shipping"]["from_shipment2"])
                if shipment2 > from_shipment2:
                    print(f'## stock added on shipping-from_shipment2')
                    retail["stock_info"]["shipping"]["from_shipment2"] = from_shipment2 + 1
                    tobe_moved -= 1
                    continue
                
                shipment3, from_shipment3 = convert_to_zero(retail["stock_info"]["shipping"]["shipment3"]), convert_to_zero(retail["stock_info"]["shipping"]["from_shipment3"])
                if shipment3 > from_shipment3:
                    print(f'## stock added on shipping-from_shipment3')
                    retail["stock_info"]["shipping"]["from_shipment3"] = from_shipment3 + 1
                    tobe_moved -= 1
                    continue

                # JOY_CONFIRM
                if total_available != (KRD4_1001 + shipment1 + shipment2 + shipment3):
                    print(f'## joy confirm needed')
                    retail_info[ref_no]["result"] = JOY_CONFIRM
                    break
                
            print(f'### {ref_no} stock reallocation finished\n')

        # 엑셀 반환
        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active

        column_names = ["ref_no", "result"
                      , "rdc-from KRD4/1001", "rdc-from Shipment 1", "rdc-from Shipment 2", "rdc-from Shipment 3"
                      , "6400", "6401", "6402", "6403", "6404", "6405", "6407", "6408", "6409", "6410", "6411"]
        new_ws.append(column_names)

        success_cnt, pass_cnt, jinny_think_cnt, joy_confirm_cnt, no_6M_sales_cnt = 0, 0, 0, 0, 0
        for ref_no, retail in retail_info.items():
            result = retail.get("result")
            if result == PASS:
                pass_cnt += 1
            if result == SUCCESS:
                success_cnt += 1
            if result == JINNY_THINK:
                jinny_think_cnt += 1
            if result == JOY_CONFIRM:
                joy_confirm_cnt += 1
            if result == NO_6M_SALES:
                no_6M_sales_cnt += 1

            stock_info, store_info = retail.get("stock_info"), retail.get("store_info")
            new_ws.append([ref_no, retail.get("result"), stock_info.get("warehouse").get("from_KRD4/1001", 0)
                          , stock_info.get("shipping").get("from_shipment1", 0), stock_info.get("shipping").get("from_shipment2", 0), stock_info.get("shipping").get("from_shipment3", 0)
                          , store_info.get("6400").get("reallocation", 0), store_info.get("6401").get("reallocation", 0)
                          , store_info.get("6402").get("reallocation", 0), store_info.get("6403").get("reallocation", 0)
                          , store_info.get("6404").get("reallocation", 0), store_info.get("6405").get("reallocation", 0)
                          , store_info.get("6407").get("reallocation", 0), store_info.get("6408").get("reallocation", 0)
                          , store_info.get("6409").get("reallocation", 0), store_info.get("6410").get("reallocation", 0)
                          , store_info.get("6411").get("reallocation", 0)])
            
        new_wb.save(filename=f'{SOURCE_DIR}/{SOURCE_FILE_NAME}_reallocation_{get_current_time("%Y%m%d_%H%M%S")}.xlsx')
        print(f"\nExcel successfully created!")
        
        total_items = len(retail_info.items())
        print(f"\n* Total: {total_items}\n- PASS: {pass_cnt}({round(pass_cnt/total_items*100, 2)}%)\n- SUCCESS: {success_cnt}({round(success_cnt/total_items*100, 2)})%\n- NO_6M_SALES: {no_6M_sales_cnt}({round(no_6M_sales_cnt/total_items*100, 2)})%\n- JINNY_THINK: {jinny_think_cnt}({round(jinny_think_cnt/total_items*100, 2)})%\n- JOY_CONFIRM: {joy_confirm_cnt}({round(joy_confirm_cnt/total_items*100, 2)})%")
        print("")

        wb.close()
        new_wb.close()

if __name__ == "__main__":
    m = Main()
    m.run()