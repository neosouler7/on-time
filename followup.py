
from collections import Counter, OrderedDict

import openpyxl

from common import STORE_MAP, PASS, SUCCESS, JINNY_THINK, JOY_CONFIRM, NO_6M_SALES
from common import get_current_time, get_column_number, get_last_data_idx, convert_to_zero

SOURCE_DIR = "C:/Users/jinny.hur/Desktop/업무파일/01. 리얼로케이션" # 마스터 파일 저장 경로
SOURCE_FILE_NAME = "Retail BTQ reallocation (0207_rank)_JH2" # 마스터 파일명(확장자 제외)


def get_retail_info_template(store_info_list):
    store_keys_list = list(STORE_MAP.keys())

    store_info = {
        store_keys_list[i]: {
            "L12M_sales": convert_to_zero(store_info_list[i * 10]),
            "L6M_sales": convert_to_zero(store_info_list[i * 10 + 1]),
            "L3M_sales": convert_to_zero(store_info_list[i * 10 + 2]),
            "MS": convert_to_zero(store_info_list[i * 10 + 3]),
            "stock": convert_to_zero(store_info_list[i * 10 + 4]),
            "in_transit": convert_to_zero(store_info_list[i * 10 + 5]),
            "wish_list": convert_to_zero(store_info_list[i * 10 + 6]),
            "reallocation": convert_to_zero(store_info_list[i * 10 + 7]),
            "coverage": convert_to_zero(store_info_list[i * 10 + 8]),
            "rotation": store_info_list[i * 10 + 9], # rotation의 경우 수식값으로 None과 0.0 이 명확하게 구분되는 바 filter 적용하지 않음
        } for i in range(len(store_keys_list))
    }

    ranking = {
        "L12M_label": "",
        "L12M_sales_sum": 0,
        "L6M_label": "",
        "L6M_sales_sum": 0,
        "L3M_label": "",
        "L3M_sales_sum": 0
    }

    return {"model": "", "entry": "", "function": "", "shortage_store_count": 0, "store_info": store_info, "ranking": ranking}

def get_product_detail(ref_no):
    if len(ref_no) >= 8 and ref_no[3] == "4":
        model = ref_no[:8] + "00"
    elif len(ref_no) >= 8 and ref_no[3] == "6":
        model = ref_no[:8] + "00"
    else:
        model = ref_no

    if len(ref_no) >= 3 and ref_no[2] == "8":
        entry = "BIJOUX"
    elif len(ref_no) >= 3 and ref_no[2] == "B":
        entry = "BIJOUX"
    elif len(ref_no) >= 3 and ref_no[2] == "N":
        entry = "NJ"
    else:
        entry = "X"

    if len(ref_no) >= 4 and ref_no[3] == "4":
        function = "RING"
    elif len(ref_no) >= 4 and ref_no[3] == "6":
        function = "BRAC"
    elif len(ref_no) >= 4 and ref_no[3] == "7":
        function = "NECK"
    elif len(ref_no) >= 4 and ref_no[3] == "8":
        function = "EAR"
    elif len(ref_no) >= 3 and ref_no[2] == "8":
        function = "EAR"
    elif len(ref_no) >= 4 and ref_no[3] == "3":
        function = "NECK"
    else:
        function = "X"

    return model, entry, function

def get_sales_sum(store_info):
    L12M_sales_sum, L6M_sales_sum, L3M_sales_sum = 0.0, 0.0, 0.0
    for store_id, info in store_info.items():
        L12M_sales_sum += info["L12M_sales"]
        L6M_sales_sum += info["L6M_sales"]
        L3M_sales_sum += info["L3M_sales"]
    return L12M_sales_sum, L6M_sales_sum, L3M_sales_sum


class Main:
    def __init__(self):
        pass
    
    def run(self):
        """
            reallocation.py 내 포함하려고 하였으나, JINNY_THINK / JOY_CONFIRM 과 같은 작업 이후에 rotation 값이 변경 되므로, 부득이하게 별도 파일 생성하게 되었음.
            즉, 1. reallocation.py 실행 / 2. JINNY 엑셀 작업 / 3. followup.py 와 같은 순서로 진행하면 됨.

            followup.py 는 크게 아래의 기능을 가짐
            1. shortage_store_count
                a. 품번 기준, rotation 값이 0인 store 개수를 저장
            2. ranking
                a. dddd
                b. ddd
                
            3. 엑셀 반환

            # data_set
            {
                "CRN7413800": {
                    "model": a,
                    "entry": a,
                    "function": a,
                    "shortage_store_count": a,
                    "store_info": {
                        "6400: {
                            "L12M_sales": a,
                            "L6M_sales": a,
                            "L3M sales": b,
                            "MS": c,
                            "stock": d,
                            "in_transit": e,
                            "wish_list": f,
                            "reallocation": g,
                            "coverage": h,
                            "rotation": i,
                        },
                        ...
                    }
                    "ranking: {
                        "L12M_label": b,
                        "L12M_sales_sum": a,
                        "L6M_label": b,
                        "L6M_sales_sum": a,
                        "L3M_label": b,
                        "L3M_sales_sum": a,
                    }
                },
                ...
            }
        """

        # 마스터 파일의 "Allocation" 시트를 참조하여, 위 형태의 data-set을 구성
        file_name = f'{SOURCE_DIR}/{SOURCE_FILE_NAME}.xlsx'
        print(f"Reading {file_name} ...")
        wb = openpyxl.load_workbook(filename=file_name, data_only=True)
        ws = wb["Allocation"]

        retail_info = dict()
        for row in ws.iter_rows(min_row=8, max_row=get_last_data_idx(ws), values_only=True):
            temp = get_retail_info_template(list(row)[get_column_number("AK"):get_column_number("EP") + 1]) # store_info
            retail_info[row[1]] = temp

        # retail_info에 필요한 정보 저장
        print("----- DATA GATHERING START -----\n")
        for ref_no, retail in retail_info.items():
            # 품번 기준, rotation 값이 0인 store 개수를 저장한다.
            shortage_store_count = 0
            for item in retail.get("store_info").items():
                if item[1].get("rotation") == 0.0:
                    shortage_store_count += 1
            retail["shortage_store_count"] = shortage_store_count

            model, entry, function = get_product_detail(ref_no)

            L12M_sales_sum, L6M_sales_sum, L3M_sales_sum = get_sales_sum(retail.get("store_info"))

            retail_info[ref_no]["model"] = model
            retail_info[ref_no]["entry"] = entry
            retail_info[ref_no]["function"] = function
            retail_info[ref_no]["ranking"]["L12M_sales_sum"] = L12M_sales_sum
            retail_info[ref_no]["ranking"]["L6M_sales_sum"] = L6M_sales_sum
            retail_info[ref_no]["ranking"]["L3M_sales_sum"] = L3M_sales_sum
            

        print("----- DATA GATHERING END -----\n")
        
        print("----- STATISTICS START -----\n")
        from operator import itemgetter

        # NJ ranking

        # for a, b in retail_info.items():
        #     print(a, b)

        for period in ["12M"]:
            nj_data = {model: {"model": product_info["model"], "sales_value": product_info["ranking"][f"L{period}_sales_sum"]} for _, product_info in retail_info.items() if product_info["entry"] == "NJ"}
            # print(nj_data)
            sorted_nj_data = sorted(nj_data.values(), key=itemgetter("sales_value"), reverse=True)[:10]
            print("sorted_nj_data")
            print(sorted_nj_data)
            print("")

            nj_ranking = {}
            for i, nj_info in enumerate(sorted_nj_data, start=1):
                print(i, nj_info)
                nj_model = nj_info["model"]
                nj_ranking[f"NJ{i}"] = {"rank": i, "sales_value": nj_info["sales_value"], "model": nj_model}

            print("nj_ranking")
            print(nj_ranking)
            print("")


            # NJ 엔트리의 순위를 기존 데이터에 적용
            for nj_rank, nj_info in nj_ranking.items():
                print(nj_rank, nj_info)
                nj_model_id = nj_info["product_id"]
                nj_model = retail_info[nj_product_id]["model"]

                print(nj_rank, nj_product_id, nj_model)
                # retail_info[nj_product_id]["ranking"]["L6M_label"] = f"{nj_model}_{nj_rank}"


        
        print("----- STATISTICS END -----\n")





        #     # reallocation과 무관하지만, 업무 필요 사항으로 ranking 추가함

        #     total_available = retail.get("stock_info").get("total_available")
        #     current_available = total_available # reallocation 위해 재고값 복사

        #     print(f'\n# START {ref_no} with {total_available} stock')

        #     if total_available == 0.0:
        #         print(f'{ref_no} pass since no TTL Available')
        #         retail_info[ref_no]["result"] = PASS
        #         continue

        #     while True:
        #         if current_available == 0:
        #             print(f'### store reallocation finished\n')
        #             break

        #         # 유효한 모든 지점의 rotation 값이 TARGET_ROTATION 초과 시, SUCCESS
        #         if all(item[1].get("rotation") > TARGET_ROTATION for item in retail["store_info"].items() if item[1].get("rotation") is not None):
        #             print(f'### {SUCCESS}\n- all store rotation over {TARGET_ROTATION} or all rotation None\n')
        #             retail_info[ref_no]["result"] = SUCCESS
        #             # time.sleep(0.5)
        #             break

        #         # L6M_sales 값이 존재하는 store 정보만 가져온다. (엑셀에서도 없을 시 제외하고 있음)
        #         store_info_items = [item for item in retail["store_info"].items() if item[1].get("L6M_sales") > 0.0]
        #         if len(store_info_items) == 0: # 전 지점에서 6개월 판매값이 없을 시
        #             print(f'### {NO_6M_SALES}\n- no L6M_sales in all stores\n')
        #             retail_info[ref_no]["result"] = NO_6M_SALES
        #             # time.sleep(0.5)
        #             break

        #         sorted_dict = OrderedDict(sorted(store_info_items, key=lambda x: x[1].get("rotation", 0) if isinstance(x[1], dict) else 0))
        #         rotation_counts = Counter(value.get("rotation", 0) if isinstance(value, dict) and isinstance(value.get("rotation"), (int, float)) else 0 for key, value in sorted_dict.items())
        #         rotation_counts = OrderedDict(sorted(rotation_counts.items()))

        #         lowest_rotation_count = rotation_counts[list(rotation_counts.keys())[0]]
        #         if lowest_rotation_count > current_available: # 현 재고에 대하여 더 이상 지점별로 동등하게 배분 못할 시
        #             print(f'### {JINNY_THINK}\n- {rotation_counts} > {current_available}\n')
        #             retail_info[ref_no]["result"] = JINNY_THINK
        #             # time.sleep(0.5)
        #             break

        #         sorted_list = sorted(store_info_items, key=lambda x: x[1].get("rotation", 0))
        #         store_id, store_info = sorted_list[0] # 가장 rotation이 낮은 store을 대상으로 함

        #         current_reallocation = store_info["reallocation"]
        #         current_rotation = store_info["rotation"]

        #         retail_info[ref_no]["store_info"][store_id]["reallocation"] += 1
        #         retail_info[ref_no]["store_info"][store_id]["rotation"] = (store_info["stock"] + store_info["in_transit"] + (current_reallocation + 1)) / (store_info["L6M_sales"] / 6)
        #         retail_info[ref_no]["result"] = SUCCESS
        #         retail_info[ref_no]["stock_info"]["tobe_moved"] += 1 # 실제 reallocation 대상 ++

        #         current_available -= 1 # reallocation 가능 재고 차감

        #         msg = f'## {STORE_MAP.get(store_id)}({store_id})\n'
        #         msg += f'- reallocation: {current_reallocation} → {retail_info[ref_no]["store_info"][store_id]["reallocation"]}\n'
        #         msg += f'- rotation: {current_rotation} → {retail_info[ref_no]["store_info"][store_id]["rotation"]}'
                
        #         print(msg)
        #         print("")
        #         # time.sleep(3)

        # # statistics & reallocation for stock_info
        # print("----- STOCK_INFO REALLOCATION -----\n")
        # for ref_no, retail in retail_info.items():


        #     if retail.get("result") not in [SUCCESS]: # SUCCESS 아니면, 재고정보 변경하지 않는다 
        #         continue

        #     total_available = retail.get("stock_info").get("total_available")
        #     tobe_moved = retail.get("stock_info").get("tobe_moved")
        #     print(f'\n# START {ref_no} with {tobe_moved} target')

        #     while tobe_moved > 0:
        #         KRD4_1001, from_KRD4_1001 = convert_to_zero(retail["stock_info"]["warehouse"]["KRD4/1001"]), convert_to_zero(retail["stock_info"]["warehouse"]["from_KRD4/1001"])
        #         if KRD4_1001 > from_KRD4_1001:
        #             print(f'## stock added on warehouse-from_KRD4/1001')
        #             retail["stock_info"]["warehouse"]["from_KRD4/1001"] = from_KRD4_1001 + 1
        #             tobe_moved -= 1
        #             continue

        #         shipment1, from_shipment1 = convert_to_zero(retail["stock_info"]["shipping"]["shipment1"]), convert_to_zero(retail["stock_info"]["shipping"]["from_shipment1"])
        #         if shipment1 > from_shipment1:
        #             print(f'## stock added on shipping-from_shipment1')
        #             retail["stock_info"]["shipping"]["from_shipment1"] = from_shipment1 + 1
        #             tobe_moved -= 1
        #             continue
                
        #         shipment2, from_shipment2 = convert_to_zero(retail["stock_info"]["shipping"]["shipment2"]), convert_to_zero(retail["stock_info"]["shipping"]["from_shipment2"])
        #         if shipment2 > from_shipment2:
        #             print(f'## stock added on shipping-from_shipment2')
        #             retail["stock_info"]["shipping"]["from_shipment2"] = from_shipment2 + 1
        #             tobe_moved -= 1
        #             continue
                
        #         shipment3, from_shipment3 = convert_to_zero(retail["stock_info"]["shipping"]["shipment3"]), convert_to_zero(retail["stock_info"]["shipping"]["from_shipment3"])
        #         if shipment3 > from_shipment3:
        #             print(f'## stock added on shipping-from_shipment3')
        #             retail["stock_info"]["shipping"]["from_shipment3"] = from_shipment3 + 1
        #             tobe_moved -= 1
        #             continue

        #         # JOY_CONFIRM
        #         if total_available != (KRD4_1001 + shipment1 + shipment2 + shipment3):
        #             print(f'## joy confirm needed')
        #             retail_info[ref_no]["result"] = JOY_CONFIRM
        #             break
                
        #     print(f'### {ref_no} stock reallocation finished\n')

        # 엑셀 반환
        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active

        # TODO. model, entry, function
        column_names = ["ref_no", "shortage_store_count", "L12M_label", "L12M_sales_sum", "L6M_label", "L6M_sales_sum", "L3M_label", "L3M_sales_sum"]
        new_ws.append(column_names)

        for ref_no, retail in retail_info.items():
            ranking = retail.get("ranking")
            new_ws.append([ref_no, retail.get("shortage_store_count"), ranking.get("L12M_label"), ranking.get("L12M_sales_sum"), ranking.get("L6M_label"), ranking.get("L6M_sales_sum"), ranking.get("L3M_label"), ranking.get("L3M_sales_sum")])
            
        new_wb.save(filename=f'{SOURCE_DIR}/{SOURCE_FILE_NAME}_followup_{get_current_time("%Y%m%d_%H%M%S")}.xlsx')
        print(f"\nExcel successfully created!")
        print("")

        wb.close()
        new_wb.close()

if __name__ == "__main__":
    m = Main()
    m.run()